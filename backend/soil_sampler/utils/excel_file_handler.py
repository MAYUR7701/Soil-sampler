from openpyxl import load_workbook
import os

from backend.settings import MEDIA_ROOT


def process_excel_sheet(ws):
    project_no = ws["H3"].value
    # Specify the range of rows and columns you want to iterate through
    start_row = 2  # Start from row 2
    # Your existing code to find the row_index
    row_index = 0
    end_row = ws.max_row  # Go up to the last row
    # Choose starting row to execute
    for index in range(start_row, end_row):
        row = ws[index]
        # Check if the first cell in the row is "LOCATION"
        if (
            row[0].value == "LOCATION"
            or row[0].value == "location"
            or row[0].value == "Location"
        ):
            row_index = index + 1
            break
    last_sample_no = None
    samples = []
    while row_index <= end_row:
        row = ws[row_index]
        sample_no = row[3].value
        top_depth = row[7].value
        location = row[0].value
        # Use the last non-empty sample_no
        if sample_no:
            last_sample_no = sample_no
        # If sample_no is empty, use the last non-empty sample_no
        elif not sample_no and last_sample_no:
            sample_no = last_sample_no
        if sample_no and top_depth and location:
            samples.append(
                {
                    "project_no": project_no,
                    "sample_no": sample_no,
                    "location": location,
                    "top_depth": top_depth,
                }
            )
        row_index += 1
    return samples


def process_excel_file(excel_folder_file_url):
    filepath=MEDIA_ROOT+excel_folder_file_url
    excel_folder_path = os.path.join(filepath)
    wb = load_workbook(filename=excel_folder_path, data_only=True)
    # Iterate over all sheets in the workbook
    sheet_data = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_data.append(process_excel_sheet(ws))
    return sheet_data
